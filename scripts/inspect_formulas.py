import openpyxl

# Load the workbook
file_path = 'c:/Users/shubh/Downloads/Dimentions Audit Authenticator/DimensionsMasterLatest.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False) # data_only=False keeps formulas
sheet = wb['Billing Dimensions']

# Row 3 is likely the first data row (1-based index)
# Row 1: Manual Entries...
# Row 2: Headers
# Row 3: Data
row_idx = 3

print(f"Inspecting formulas in Row {row_idx}:")

# Iterate through columns to find formulas
for col_idx, cell in enumerate(sheet[row_idx], start=1):
    val = cell.value
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Check if it starts with '=' indicating a formula
    if isinstance(val, str) and val.startswith('='):
        print(f"Column {col_letter} ({col_idx}): {val}")
    else:
        # Print value if it looks like a calculated result column (just to check)
        if col_idx > 10: # Focus on later columns where calcs usually happen
             print(f"Column {col_letter} ({col_idx}): Value='{val}'")

print("\n--- Header Map (Row 2) ---")
for col_idx, cell in enumerate(sheet[2], start=1):
    print(f"{openpyxl.utils.get_column_letter(col_idx)}: {cell.value}")
