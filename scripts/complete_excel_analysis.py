import openpyxl
import pandas as pd

# Load workbook with formulas preserved
wb = openpyxl.load_workbook('c:/Users/shubh/Downloads/Dimentions Audit Authenticator/DimensionsMasterLatest.xlsx', data_only=False)
wb_values = openpyxl.load_workbook('c:/Users/shubh/Downloads/Dimensions Audit Authenticator/DimensionsMasterLatest.xlsx', data_only=True)

sheet = wb['Billing Dimensions']
sheet_values = wb_values['Billing Dimensions']

print("="*80)
print("COMPLETE EXCEL ANALYSIS - BILLING DIMENSIONS SHEET")
print("="*80)

# Row 1: Top headers
print("\n--- ROW 1: Section Headers ---")
for col_idx in range(1, 35):
    cell = sheet.cell(1, col_idx)
    if cell.value:
        print(f"Column {openpyxl.utils.get_column_letter(col_idx)}: {cell.value}")

# Row 2: Main column headers
print("\n--- ROW 2: Column Names ---")
headers = []
for col_idx in range(1, 35):
    cell = sheet.cell(2, col_idx)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    headers.append((col_letter, cell.value))
    print(f"{col_letter}: {cell.value}")

# Row 3: Units/Sub-headers
print("\n--- ROW 3: Units/Sub-Headers ---")
for col_idx in range(1, 35):
    cell = sheet.cell(3, col_idx)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"{col_letter}: {cell.value}")

# Row 4: First data row - FORMULAS
print("\n" + "="*80)
print("ROW 4 (FIRST DATA ROW) - FORMULAS & VALUES")
print("="*80)
for col_idx in range(1, 35):
    cell_formula = sheet.cell(4, col_idx)
    cell_value = sheet_values.cell(4, col_idx)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    header = headers[col_idx-1][1]
    
    if isinstance(cell_formula.value, str) and cell_formula.value.startswith('='):
        print(f"\n{col_letter} ({header}):")
        print(f"  FORMULA: {cell_formula.value}")
        print(f"  VALUE: {cell_value.value}")
    else:
        print(f"\n{col_letter} ({header}): {cell_value.value}")

# Identify weight-related columns
print("\n" + "="*80)
print("WEIGHT TYPE IDENTIFICATION")
print("="*80)
weight_keywords = ['weight', 'wht', 'kg', 'gm', 'vol', 'cbf', 'chargeable', 'bom', 'theoretical']
for col_letter, header in headers:
    if header and any(keyword in str(header).lower() for keyword in weight_keywords):
        print(f"{col_letter}: {header}")

# Check for Parent/Child relationship columns
print("\n" + "="*80)
print("PARENT/CHILD RELATIONSHIP COLUMNS")
print("="*80)
relationship_keywords = ['mtp', 'sku', 'parent', 'child', 'sb', 'mb']
for col_letter, header in headers:
    if header and any(keyword in str(header).lower() for keyword in relationship_keywords):
        print(f"{col_letter}: {header}")

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
