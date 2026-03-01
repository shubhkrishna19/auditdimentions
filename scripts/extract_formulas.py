import openpyxl

# Load workbook with formulas (data_only=False)
wb = openpyxl.load_workbook('DimensionsMasterLatest.xlsx', data_only=False)
sheet = wb['Billing Dimensions']

print("="*100)
print("FORMULA ANALYSIS - BILLING DIMENSIONS")
print("="*100)

# Analyze rows 4, 5, 6 to see formula patterns
for row_num in [4, 5, 6]:
    print(f"\n{'='*100}")
    print(f"ROW {row_num} FORMULAS")
    print(f"{'='*100}")
    
    sku = sheet.cell(row_num, 1).value
    print(f"SKU: {sku}\n")
    
    # Check columns C through AE for formulas
    for col_idx in range(3, 32):  # C to AE
        cell = sheet.cell(row_num, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        if isinstance(cell.value, str) and cell.value.startswith('='):
            # Get header from row 2
            header = sheet.cell(2, col_idx).value
            unit = sheet.cell(3, col_idx).value
            
            print(f"\n{col_letter} - {header} ({unit}):")
            print(f"  Formula: {cell.value}")

# Now let's specifically map the weight calculation columns
print("\n" + "="*100)
print("WEIGHT CALCULATION COLUMN MAPPING")
print("="*100)

weight_cols = {
    'O': 'CBF/Volume Calc',
    'P': 'Column labeled "6" (Kgs)',
    'Q': 'Phys.Wht (Gms)',
    'R': 'Wht(Kgs)',
}

row = 4
print(f"\nAnalyzing Row {row}:")
for col_letter, description in weight_cols.items():
    col_idx = openpyxl.utils.get_column_letter_index(col_letter)
    cell = sheet.cell(row, col_idx)
    header = sheet.cell(2, col_idx).value
    
    print(f"\n{col_letter} - {description} (Header: {header}):")
    if isinstance(cell.value, str) and cell.value.startswith('='):
        print(f"  FORMULA: {cell.value}")
    else:
        print(f"  VALUE: {cell.value}")

# Check if there's a BOM/Theoretical weight column
print("\n" + "="*100)
print("SEARCHING FOR BOM/THEORETICAL WEIGHT")
print("="*100)

for col_idx in range(1, 35):
    header = sheet.cell(2, col_idx).value
    if header and ('bom' in str(header).lower() or 'theoretical' in str(header).lower() or 'theor' in str(header).lower()):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"Found at {col_letter}: {header}")
        
        # Check formula
        cell = sheet.cell(4, col_idx)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  Formula: {cell.value}")

print("\n" + "="*100)
print("ANALYSIS COMPLETE")
print("="*100)
